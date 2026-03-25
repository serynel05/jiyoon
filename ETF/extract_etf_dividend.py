import re
from pathlib import Path
import pandas as pd
from typing import Tuple

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet


def parse_kakao_pension_txt(text: str) -> pd.DataFrame:
    """
    카톡 TXT에서 다음 형태를 추출:
    ■ ETF명 : KODEX 은행
    ■ 입금액 : 450원
    ■ 입금일 : 2026.02.03
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    records = []
    cur = {"ETF명": None, "입금액(원)": None, "입금일": None}

    etf_pat = re.compile(r"^■\s*ETF명\s*:\s*(.+?)\s*$")
    amt_pat = re.compile(r"^■\s*입금액\s*:\s*([0-9,]+)\s*원?\s*$")
    date_pat = re.compile(r"^■\s*입금일\s*:\s*([0-9]{4})[.\-/]([0-9]{1,2})[.\-/]([0-9]{1,2})\s*$")

    def flush_if_complete():
        nonlocal cur
        if cur["ETF명"] and cur["입금액(원)"] is not None and cur["입금일"] is not None:
            records.append(cur)
        cur = {"ETF명": None, "입금액(원)": None, "입금일": None}

    for ln in lines:
        m = etf_pat.match(ln)
        if m:
            flush_if_complete()
            cur["ETF명"] = m.group(1).strip()
            continue

        m = amt_pat.match(ln)
        if m:
            cur["입금액(원)"] = int(m.group(1).replace(",", ""))
            continue

        m = date_pat.match(ln)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            cur["입금일"] = pd.Timestamp(year=y, month=mo, day=d)
            continue

    flush_if_complete()

    df = pd.DataFrame(records)
    if df.empty:
        return df

    df["월"] = df["입금일"].dt.to_period("M").astype(str)  # "YYYY-MM"
    return df


def build_reports(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    return:
      detail: 상세
      summary_month_etf: 월xETF (long)
      pivot_month_etf: 월행 / ETF열 (wide)
      summary_etf: ETF별 전체 합
    """
    if df.empty:
        detail = pd.DataFrame(columns=["월", "입금일", "ETF명", "입금액(원)"])
        summary_month_etf = pd.DataFrame(columns=["월", "ETF명", "총 입금액(원)", "건수"])
        pivot_month_etf = pd.DataFrame()
        summary_etf = pd.DataFrame(columns=["ETF명", "총 입금액(원)", "건수"])
        return detail, summary_month_etf, pivot_month_etf, summary_etf

    detail = df[["월", "입금일", "ETF명", "입금액(원)"]].sort_values(["입금일", "ETF명"]).reset_index(drop=True)

    summary_month_etf = (
        df.groupby(["월", "ETF명"], as_index=False)
          .agg(**{"총 입금액(원)": ("입금액(원)", "sum"),
                 "건수": ("입금액(원)", "count")})
          .sort_values(["월", "총 입금액(원)", "ETF명"], ascending=[True, False, True])
          .reset_index(drop=True)
    )

    pivot_month_etf = (
        df.pivot_table(index="월", columns="ETF명", values="입금액(원)", aggfunc="sum", fill_value=0)
          .sort_index()
    )
    pivot_month_etf.loc[:, "월 합계"] = pivot_month_etf.sum(axis=1)
    pivot_month_etf = pivot_month_etf.reset_index()

    summary_etf = (
        df.groupby("ETF명", as_index=False)
          .agg(**{"총 입금액(원)": ("입금액(원)", "sum"),
                 "건수": ("입금액(원)", "count")})
          .sort_values(["총 입금액(원)", "ETF명"], ascending=[False, True])
          .reset_index(drop=True)
    )

    return detail, summary_month_etf, pivot_month_etf, summary_etf


def format_number_columns(ws: Worksheet, col_indices, start_row: int, end_row: int, fmt: str = "#,##0"):
    for col_idx in col_indices:
        col_letter = get_column_letter(col_idx)
        for r in range(start_row, end_row + 1):
            ws[f"{col_letter}{r}"].number_format = fmt


def append_total_row(ws: Worksheet, label_col: int, sum_col: int, start_data_row: int, end_data_row: int):
    total_row = end_data_row + 1

    ws.cell(row=total_row, column=label_col, value="총합")
    ws.cell(row=total_row, column=label_col).font = Font(bold=True)
    ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal="right")

    sum_col_letter = get_column_letter(sum_col)
    ws.cell(
        row=total_row,
        column=sum_col,
        value=f"=SUM({sum_col_letter}{start_data_row}:{sum_col_letter}{end_data_row})"
    )
    ws.cell(row=total_row, column=sum_col).font = Font(bold=True)
    ws.cell(row=total_row, column=sum_col).number_format = "#,##0"


def main(input_txt_path: str, output_xlsx_path: str = "output.xlsx"):
    text = Path(input_txt_path).read_text(encoding="utf-8")
    df = parse_kakao_pension_txt(text)

    detail, summary_month_etf, pivot_month_etf, summary_etf = build_reports(df)

    with pd.ExcelWriter(output_xlsx_path, engine="openpyxl") as writer:
        summary_month_etf.to_excel(writer, index=False, sheet_name="요약(월xETF)")
        pivot_month_etf.to_excel(writer, index=False, sheet_name="피벗(월행-ETF열)")
        summary_etf.to_excel(writer, index=False, sheet_name="요약(ETF별)")
        detail.to_excel(writer, index=False, sheet_name="상세(추출원문)")

        wb = writer.book

        # 요약(월xETF)
        ws1 = wb["요약(월xETF)"]
        s, e = 2, ws1.max_row
        if e >= s:
            format_number_columns(ws1, [3], s, e, "#,##0")
            append_total_row(ws1, label_col=2, sum_col=3, start_data_row=s, end_data_row=e)

        # 피벗
        ws2 = wb["피벗(월행-ETF열)"]
        s, e = 2, ws2.max_row
        if e >= s and ws2.max_column >= 2:
            format_number_columns(ws2, list(range(2, ws2.max_column + 1)), s, e, "#,##0")

        # 요약(ETF별)
        ws3 = wb["요약(ETF별)"]
        s, e = 2, ws3.max_row
        if e >= s:
            format_number_columns(ws3, [2], s, e, "#,##0")
            append_total_row(ws3, label_col=1, sum_col=2, start_data_row=s, end_data_row=e)

        # 상세
        ws4 = wb["상세(추출원문)"]
        s, e = 2, ws4.max_row
        if e >= s:
            format_number_columns(ws4, [4], s, e, "#,##0")

    print(f"완료: {output_xlsx_path}")
    if not detail.empty:
        print(f"- 추출 {len(detail)}건, ETF {detail['ETF명'].nunique()}종, 월 {detail['월'].nunique()}개")
    else:
        print("- 추출 0건")


if __name__ == "__main__":
    import sys

    # 사용법:
    # 1) python kakao_pension_monthly.py input.txt        -> output.xlsx
    # 2) python kakao_pension_monthly.py input.txt out.xlsx -> out.xlsx
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("사용법: python kakao_pension_monthly.py <input.txt> [output.xlsx]")
        raise SystemExit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) == 3 else "output.xlsx"
    main(input_path, output_path)